from flask import Flask, request, jsonify, send_file, render_template
import openpyxl
from openpyxl.styles import Font, Border, Alignment, PatternFill, Side
import os
import json
import io
import copy
from datetime import datetime, date
import sys

def resource_path(relative_path):
    """Obtiene la ruta absoluta al recurso, compatible con PyInstaller y desarrollo."""
    if getattr(sys, 'frozen', False):
        return os.path.join(sys._MEIPASS, relative_path)
    return os.path.join(os.path.abspath("."), relative_path)

# Para archivos modificables (Excel, JSON), buscar en la carpeta donde está el ejecutable (o el script)
base_dir = os.path.dirname(sys.executable) if getattr(sys, 'frozen', False) else os.path.dirname(os.path.abspath(__file__))

app = Flask(
    __name__,
    template_folder=resource_path('templates'),
    static_folder=resource_path('static')
)

DNI_FILE = os.path.join(base_dir, 'vendors_dni.json')
MASTER_FILE = os.path.join(base_dir, '2. PASAJES FFVV JULIO 2026 MODELO.xlsx')
TEMPLATE_FILE = os.path.join(base_dir, '1. Formato Planilla de movilidad - por trabajador.xlsx')


# Bancos válidos (hojas en el archivo maestro)
BANCOS_VALIDOS = ['BBVA', 'BCP', 'SCOTIABANK', 'FINANCIERO']

# Filas fijas por bloque de vendedor en el archivo maestro
BLOCK_SIZE = 26

# ──────────────────────────────────────────────────────────
# Utilidades de persistencia
# ──────────────────────────────────────────────────────────

def load_dnis():
    """Carga el mapeo vendedor→DNI y banco desde JSON local."""
    if os.path.exists(DNI_FILE):
        try:
            with open(DNI_FILE, 'r', encoding='utf-8') as f:
                return json.load(f)
        except Exception:
            pass
    return {}

def save_vendor_info(vendor_name, dni, banco):
    """Guarda o actualiza el DNI y banco de un vendedor."""
    dnis = load_dnis()
    dnis[vendor_name.strip()] = {
        'dni': dni.strip(),
        'banco': banco.strip().upper()
    }
    try:
        with open(DNI_FILE, 'w', encoding='utf-8') as f:
            json.dump(dnis, f, ensure_ascii=False, indent=4)
    except Exception as e:
        print("Error saving vendor JSON:", e)

# ──────────────────────────────────────────────────────────
# Extracción de datos del archivo maestro
# ──────────────────────────────────────────────────────────

def extract_vendors_and_defaults():
    """
    Lee el archivo maestro (JULIO 2026 MODELO).
    Estructura por banco (hoja) → vendedores → rutas predeterminadas por día de semana.
    Columnas en el modelo: B=VENDEDOR, C=DIA, D=RUTA, E=LUGAR, F=MONTO, G=TOTAL
    Retorna dict: { "NOMBRE VENDEDOR": { "banco": "BCP", "defaults": { "LUNES": {...} } } }
    """
    vendors = {}
    if not os.path.exists(MASTER_FILE):
        return vendors

    try:
        wb = openpyxl.load_workbook(MASTER_FILE, data_only=True)
        for sheet_name in wb.sheetnames:
            # Solo procesamos hojas de bancos
            if sheet_name.upper() not in [b.upper() for b in BANCOS_VALIDOS]:
                continue

            sheet = wb[sheet_name]
            banco_nombre = sheet_name.upper()

            current_vendor = None
            for r in range(3, sheet.max_row + 1):
                val_b = sheet.cell(row=r, column=2).value  # VENDEDOR
                val_c = sheet.cell(row=r, column=3).value  # DIA
                val_f = sheet.cell(row=r, column=6).value  # MONTO

                # Fin de tabla en fila TOTAL
                if val_b and str(val_b).strip().upper() == 'TOTAL':
                    break

                # Nueva entrada de vendedor
                if val_b is not None and str(val_b).strip():
                    current_vendor = str(val_b).strip()
                    if current_vendor not in vendors:
                        vendors[current_vendor] = {
                            "banco": banco_nombre,
                            "defaults": {}
                        }

                if current_vendor:
                    dia = str(val_c).strip().upper() if val_c else ""
                    ruta = sheet.cell(row=r, column=4).value   # col D
                    lugar = sheet.cell(row=r, column=5).value  # col E
                    monto = sheet.cell(row=r, column=6).value  # col F

                    if dia in ['LUNES', 'MARTES', 'MIERCOLES', 'JUEVES', 'VIERNES', 'SABADO']:
                        # Solo la primera ocurrencia de cada día (sin duplicados)
                        if dia not in vendors[current_vendor]["defaults"]:
                            vendors[current_vendor]["defaults"][dia] = {
                                "ruta": str(ruta).strip() if ruta else "",
                                "lugar": str(lugar).strip() if lugar else "",
                                "monto": float(monto) if monto is not None else 0.0
                            }
        wb.close()
    except Exception as e:
        print("Error extracting vendor defaults:", e)

    return vendors

# ──────────────────────────────────────────────────────────
# Utilidades de estilo
# ──────────────────────────────────────────────────────────

def copy_cell_style(source_cell, target_cell):
    if source_cell.has_style:
        target_cell.font = copy.copy(source_cell.font)
        target_cell.border = copy.copy(source_cell.border)
        target_cell.fill = copy.copy(source_cell.fill)
        target_cell.alignment = copy.copy(source_cell.alignment)
        target_cell.number_format = source_cell.number_format

def make_thin_border():
    s = Side(style='thin', color='000000')
    return Border(left=s, right=s, top=s, bottom=s)

def make_ref_font():
    return Font(name="Calibri", size=10)

# ──────────────────────────────────────────────────────────
# Rutas Flask
# ──────────────────────────────────────────────────────────

@app.route('/')
def index():
    return render_template('index.html')

@app.route('/rrhh')
def rrhh():
    return render_template('rrhh.html')

@app.route('/api/get-vendors', methods=['GET'])
def get_vendors():
    vendors = extract_vendors_and_defaults()
    stored = load_dnis()
    # Enriquecer con DNI guardado
    dnis = {}
    bancos = {}
    for name, info in stored.items():
        if isinstance(info, dict):
            dnis[name] = info.get('dni', '')
            bancos[name] = info.get('banco', '')
        else:
            # Compatibilidad con formato antiguo (solo DNI como string)
            dnis[name] = str(info)
    # Si el banco ya está en vendors del archivo maestro, lo actualizamos
    for name, vdata in vendors.items():
        if name not in bancos:
            bancos[name] = vdata.get('banco', '')
    return jsonify({
        "vendors": vendors,
        "dnis": dnis,
        "bancos": bancos
    })

@app.route('/api/generate-planilla', methods=['POST'])
def generate_planilla():
    data = request.get_json()
    if not data:
        return jsonify({"status": "error", "message": "No data received"}), 400

    vendor          = data.get('vendedor', '').strip()
    dni             = data.get('dni', '').strip()
    banco           = data.get('banco', '').strip().upper()
    periodo_str     = data.get('periodo', '').strip()      # e.g. "2026-07"
    fecha_emision_s = data.get('fecha_emision', '').strip() # e.g. "2026-07-31"
    planilla_nro    = int(data.get('planilla_nro_inicial', 1))
    rows            = data.get('rows', [])

    if not vendor or not dni or not banco or not periodo_str or not fecha_emision_s or not rows:
        return jsonify({"status": "error", "message": "Faltan campos obligatorios"}), 400

    # Persistir info del vendedor
    save_vendor_info(vendor, dni, banco)

    if not os.path.exists(TEMPLATE_FILE):
        return jsonify({"status": "error", "message": "Archivo plantilla no encontrado"}), 500

    try:
        wb = openpyxl.load_workbook(TEMPLATE_FILE)
        template_sheet = wb['Planilla Movilidad']

        p_yr, p_mo = map(int, periodo_str.split('-'))
        periodo_date   = date(p_yr, p_mo, 1)
        fecha_emision  = datetime.strptime(fecha_emision_s, "%Y-%m-%d").date()

        # Paginación: 14 filas por hoja de planilla
        chunk_size = 14
        row_chunks = [rows[i:i + chunk_size] for i in range(0, len(rows), chunk_size)]

        for p_idx, chunk in enumerate(row_chunks):
            if p_idx == 0:
                sheet = template_sheet
                sheet.title = f"Planilla Pg {p_idx + 1}"
            else:
                sheet = wb.copy_worksheet(template_sheet)
                sheet.title = f"Planilla Pg {p_idx + 1}"

            # ── Cabecera ──
            sheet['F13'] = vendor
            sheet['F14'] = dni
            sheet['K11'] = banco           # ← NUEVO: Banco en K11
            sheet['D8']  = periodo_date
            sheet['D8'].number_format = 'mmm-yy'
            sheet['K5']  = fecha_emision
            sheet['K5'].number_format = 'dd/mm/yyyy'
            sheet['K1']  = f"{planilla_nro + p_idx:02d}"

            # ── Filas de desplazamiento (18–31) ──
            for r_idx in range(chunk_size):
                excel_row = 18 + r_idx
                if r_idx < len(chunk):
                    rec = chunk[r_idx]
                    sheet.cell(row=excel_row, column=1,  value=int(rec['dia']))
                    sheet.cell(row=excel_row, column=2,  value=int(rec['mes']))
                    sheet.cell(row=excel_row, column=3,  value=int(rec['ano']))
                    sheet.cell(row=excel_row, column=4,  value=rec['motivo'])
                    sheet.cell(row=excel_row, column=6,  value=rec['ruta'])
                    sheet.cell(row=excel_row, column=10, value=rec['lugar'])
                    sheet.cell(row=excel_row, column=11, value=float(rec['monto']))
                else:
                    for col in [1, 2, 3, 4, 6, 10, 11]:
                        sheet.cell(row=excel_row, column=col, value=None)

            sheet['K32'] = "=SUM(K18:K31)"

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        wb.close()

        filename = (
            f"Planilla_Movilidad_{vendor.replace(' ', '_')}"
            f"_{banco}_{periodo_str.replace('-', '_')}.xlsx"
        )
        return send_file(
            output,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            as_attachment=True,
            download_name=filename
        )
    except Exception as e:
        print("Error generando planilla:", e)
        return jsonify({"status": "error", "message": str(e)}), 500

import tempfile
import uuid

# Almacen temporal de consolidados generados (token -> filepath)
_consolidado_temp = {}

@app.route('/api/generate-consolidado', methods=['POST'])
def generate_consolidado():
    """
    Paso 1: Recibe planillas, genera el Excel en un archivo temporal,
    retorna JSON con logs y un token de descarga.
    El archivo maestro en disco NUNCA se modifica.
    """
    if 'files' not in request.files:
        return jsonify({"status": "error", "message": "No se subieron archivos"}), 400

    uploaded_files = request.files.getlist('files')
    if not uploaded_files or uploaded_files[0].filename == '':
        return jsonify({"status": "error", "message": "No se seleccionó ningún archivo"}), 400

    if not os.path.exists(MASTER_FILE):
        return jsonify({"status": "error", "message": "Archivo maestro no encontrado"}), 500

    try:
        # ── Cargar estructura del maestro como base ──
        master_wb = openpyxl.load_workbook(MASTER_FILE)

        # Limpiar datos de vendedores de las hojas de banco,
        # dejando solo cabecera y fila TOTAL
        for sheet_name in master_wb.sheetnames:
            if sheet_name.upper() not in [b.upper() for b in BANCOS_VALIDOS]:
                continue
            sheet = master_wb[sheet_name]
            total_row = None
            for r in range(3, sheet.max_row + 1):
                val = sheet.cell(row=r, column=2).value
                if val and str(val).strip().upper() == 'TOTAL':
                    total_row = r
                    break
            # Eliminar filas de datos entre la fila 3 y TOTAL (exclusive)
            if total_row and total_row > 3:
                # Quitar merges en ese rango
                to_remove = [
                    rng for rng in list(sheet.merged_cells.ranges)
                    if rng.bounds[1] >= 3 and rng.bounds[3] < total_row
                ]
                for rng in to_remove:
                    sheet.merged_cells.remove(rng)
                rows_to_del = total_row - 3
                sheet.delete_rows(3, rows_to_del)

        consolidated_count = 0
        execution_logs = []
        periodo_detectado = None

        # ── Procesar cada planilla subida ──
        for file in uploaded_files:
            stream = io.BytesIO(file.read())
            worker_wb = openpyxl.load_workbook(stream, data_only=True)

            vendor_name = None
            banco_name = None
            all_records = []
            periodo_file = None

            for ws in worker_wb.worksheets:
                v_name = ws['F13'].value
                v_banco = ws['K11'].value
                v_periodo = ws['D8'].value  # Periodo (fecha)

                if v_name:
                    vendor_name = str(v_name).strip()
                if v_banco:
                    banco_name = str(v_banco).strip().upper()
                if v_periodo and periodo_file is None:
                    if isinstance(v_periodo, (datetime, date)):
                        periodo_file = v_periodo.strftime('%Y_%m')
                    else:
                        try:
                            dt = datetime.strptime(str(v_periodo).strip()[:7], '%Y-%m')
                            periodo_file = dt.strftime('%Y_%m')
                        except Exception:
                            periodo_file = str(v_periodo).strip().replace('-', '_')[:7]

                # Filas 18 a 31 (14 filas por hoja)
                for r in range(18, 32):
                    dia_val = ws.cell(row=r, column=1).value
                    if dia_val is None:
                        continue
                    try:
                        d = int(dia_val)
                        mo = int(ws.cell(row=r, column=2).value or 1)
                        yr = int(ws.cell(row=r, column=3).value or 2026)
                        date_obj = date(yr, mo, d)
                        wmap = ['LUNES', 'MARTES', 'MIERCOLES', 'JUEVES',
                                'VIERNES', 'SABADO', 'DOMINGO']
                        dia_semana = wmap[date_obj.weekday()]
                    except Exception:
                        dia_semana = "OTRO"

                    ruta = ws.cell(row=r, column=6).value
                    lugar = ws.cell(row=r, column=10).value
                    monto = ws.cell(row=r, column=11).value

                    all_records.append({
                        "dia_semana": dia_semana,
                        "ruta": str(ruta).strip() if ruta else "",
                        "lugar": str(lugar).strip() if lugar else "",
                        "monto": float(monto) if monto is not None else 0.0
                    })

            worker_wb.close()

            if periodo_file and periodo_detectado is None:
                periodo_detectado = periodo_file

            if not vendor_name:
                execution_logs.append(f"Archivo '{file.filename}' ignorado: sin nombre de trabajador.")
                continue
            if not banco_name or banco_name not in BANCOS_VALIDOS:
                execution_logs.append(
                    f"Archivo '{file.filename}' ignorado: banco '{banco_name}' no válido."
                )
                continue
            if not all_records:
                execution_logs.append(f"Archivo '{file.filename}' ignorado: sin registros de viaje.")
                continue

            # ── Seleccionar hoja del banco en el workbook de salida ──
            target_sheet_name = None
            for sname in master_wb.sheetnames:
                if sname.upper() == banco_name:
                    target_sheet_name = sname
                    break

            if not target_sheet_name:
                target_sheet_name = banco_name
                new_ws = master_wb.create_sheet(title=target_sheet_name)
                new_ws.merge_cells('B1:G1')
                new_ws['B1'] = f'TABLA DE PASAJES  - {banco_name}'
                for col, hdr in enumerate(['VENDEDOR', 'DIA ', 'RUTA', 'LUGAR', 'MONTO', 'TOTAL'], start=2):
                    new_ws.cell(row=2, column=col, value=hdr)
                new_ws.merge_cells('B3:F3')
                new_ws['B3'] = 'TOTAL'

            sheet = master_wb[target_sheet_name]

            # ── Buscar bloque existente del vendedor (por si ya fue insertado en esta sesión) ──
            start_row = None
            end_row = None
            for r in range(3, sheet.max_row + 1):
                val_b = sheet.cell(row=r, column=2).value
                if val_b and str(val_b).strip().upper() == vendor_name.upper():
                    start_row = r
                    for rng in sheet.merged_cells.ranges:
                        mc, mr, xc, xr = rng.bounds
                        if mc == 2 and mr == r:
                            end_row = xr
                            break
                    if end_row is None:
                        end_row = r
                    break

            if start_row is not None:
                to_remove = [
                    rng for rng in list(sheet.merged_cells.ranges)
                    if rng.bounds[1] >= start_row and rng.bounds[3] <= end_row
                ]
                for rng in to_remove:
                    sheet.merged_cells.remove(rng)
                rows_to_del = end_row - start_row + 1
                sheet.delete_rows(start_row, rows_to_del)
                insert_row = start_row
            else:
                total_row = None
                for r in range(3, sheet.max_row + 1):
                    v = sheet.cell(row=r, column=2).value
                    if v and str(v).strip().upper() == 'TOTAL':
                        total_row = r
                        break
                insert_row = total_row if total_row else sheet.max_row + 1

            # ── Insertar bloque del vendedor ──
            N = BLOCK_SIZE
            sheet.insert_rows(insert_row, N)

            ref_font = make_ref_font()
            ref_border = make_thin_border()

            c_vend = sheet.cell(row=insert_row, column=2, value=vendor_name)
            c_vend.font = Font(name="Calibri", size=10, bold=True)
            c_vend.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            c_vend.border = copy.copy(ref_border)
            sheet.merge_cells(
                start_row=insert_row, start_column=2,
                end_row=insert_row + N - 1, end_column=2
            )

            total_formula = f"=SUM(F{insert_row}:F{insert_row + N - 1})"
            c_total = sheet.cell(row=insert_row, column=7, value=total_formula)
            c_total.font = copy.copy(ref_font)
            c_total.alignment = Alignment(horizontal="right", vertical="center")
            c_total.number_format = '#,##0.00'
            c_total.border = copy.copy(ref_border)
            sheet.merge_cells(
                start_row=insert_row, start_column=7,
                end_row=insert_row + N - 1, end_column=7
            )

            for idx in range(N):
                curr = insert_row + idx
                for col in [2, 7]:
                    sheet.cell(row=curr, column=col).border = copy.copy(ref_border)

                if idx < len(all_records):
                    rec = all_records[idx]
                    c_dia = sheet.cell(row=curr, column=3, value=rec['dia_semana'])
                    c_ruta = sheet.cell(row=curr, column=4, value=rec['ruta'])
                    c_lugar = sheet.cell(row=curr, column=5, value=rec['lugar'])
                    c_monto = sheet.cell(row=curr, column=6, value=rec['monto'])

                    c_dia.alignment = Alignment(horizontal="center", vertical="center")
                    c_ruta.alignment = Alignment(horizontal="left", vertical="center")
                    c_lugar.alignment = Alignment(horizontal="center", vertical="center")
                    c_monto.alignment = Alignment(horizontal="right", vertical="center")
                    c_monto.number_format = '#,##0.00'

                    for col in [3, 4, 5, 6]:
                        sheet.cell(row=curr, column=col).font = copy.copy(ref_font)
                        sheet.cell(row=curr, column=col).border = copy.copy(ref_border)
                else:
                    for col in [3, 4, 5, 6]:
                        cell = sheet.cell(row=curr, column=col)
                        cell.value = None
                        cell.border = copy.copy(ref_border)

            # ── Actualizar fórmula TOTAL global de la hoja ──
            grand_total_row = None
            for r in range(3, sheet.max_row + 1):
                v = sheet.cell(row=r, column=2).value
                if v and str(v).strip().upper() == 'TOTAL':
                    grand_total_row = r
                    break

            if not grand_total_row:
                grand_total_row = sheet.max_row + 1
                sheet.cell(row=grand_total_row, column=2, value="TOTAL")
                sheet.merge_cells(
                    start_row=grand_total_row, start_column=2,
                    end_row=grand_total_row, end_column=6
                )
                for col in range(2, 7):
                    sheet.cell(row=grand_total_row, column=col).border = copy.copy(ref_border)

            sheet.cell(
                row=grand_total_row, column=7,
                value=f"=SUM(G3:G{grand_total_row - 1})"
            )
            sheet.cell(row=grand_total_row, column=7).font = Font(name="Calibri", size=10, bold=True)
            sheet.cell(row=grand_total_row, column=7).border = copy.copy(ref_border)
            sheet.cell(row=grand_total_row, column=7).alignment = Alignment(
                horizontal="right", vertical="center"
            )

            n_regs = min(len(all_records), N)
            execution_logs.append(
                f"[OK] {vendor_name} en hoja '{target_sheet_name}': "
                f"{n_regs} registros consolidados."
            )
            consolidated_count += 1

        if consolidated_count == 0:
            master_wb.close()
            return app.response_class(
                response=json.dumps({
                    "status":  "error",
                    "message": "No se pudo consolidar ningun archivo. Revisa los detalles.",
                    "logs":    execution_logs
                }, ensure_ascii=True),
                mimetype='application/json',
                status=400
            )

        # ── Guardar en archivo temporal y retornar token ──
        periodo_str = periodo_detectado or datetime.now().strftime('%Y_%m')
        filename    = f"Consolidado_Pasajes_FFVV_{periodo_str}.xlsx"

        tmp_path = os.path.join(tempfile.gettempdir(), f"consolidado_{uuid.uuid4().hex}.xlsx")
        master_wb.save(tmp_path)
        master_wb.close()

        token = uuid.uuid4().hex
        _consolidado_temp[token] = (tmp_path, filename)

        return app.response_class(
            response=json.dumps({
                "status":             "success",
                "consolidated_count": consolidated_count,
                "logs":               execution_logs,
                "token":              token,
                "filename":           filename
            }, ensure_ascii=True),
            mimetype='application/json'
        )

    except Exception as e:
        print("Error generando consolidado:", e)
        return app.response_class(
            response=json.dumps({"status": "error", "message": str(e)}, ensure_ascii=True),
            mimetype='application/json',
            status=500
        )



@app.route('/api/download-consolidado/<token>', methods=['GET'])
def download_consolidado(token):
    """Paso 2: descarga el archivo generado por token y lo borra del disco."""
    if token not in _consolidado_temp:
        return jsonify({"status": "error", "message": "Token invalido o expirado"}), 404
    tmp_path, filename = _consolidado_temp.pop(token)
    if not os.path.exists(tmp_path):
        return jsonify({"status": "error", "message": "Archivo no encontrado"}), 404
    return send_file(
        tmp_path,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name=filename
    )


@app.route('/api/consolidate', methods=['POST'])
def consolidate():
    if 'files' not in request.files:
        return jsonify({"status": "error", "message": "No se subieron archivos"}), 400

    uploaded_files = request.files.getlist('files')
    if not uploaded_files or uploaded_files[0].filename == '':
        return jsonify({"status": "error", "message": "No se seleccionó ningún archivo"}), 400

    if not os.path.exists(MASTER_FILE):
        return jsonify({"status": "error", "message": "Archivo maestro no encontrado"}), 500

    try:
        master_wb       = openpyxl.load_workbook(MASTER_FILE)
        consolidated_count = 0
        execution_logs  = []

        for file in uploaded_files:
            stream      = io.BytesIO(file.read())
            worker_wb   = openpyxl.load_workbook(stream, data_only=True)

            vendor_name = None
            banco_name  = None
            all_records = []

            for ws in worker_wb.worksheets:
                v_name  = ws['F13'].value
                v_banco = ws['K11'].value   # ← NUEVO: leer banco desde K11

                if v_name:
                    vendor_name = str(v_name).strip()
                if v_banco:
                    banco_name = str(v_banco).strip().upper()

                # Filas 18 a 31 (14 filas por hoja)
                for r in range(18, 32):
                    dia_val = ws.cell(row=r, column=1).value
                    if dia_val is None:
                        continue
                    try:
                        d  = int(dia_val)
                        mo = int(ws.cell(row=r, column=2).value or 1)
                        yr = int(ws.cell(row=r, column=3).value or 2026)
                        date_obj = date(yr, mo, d)
                        wmap = ['LUNES', 'MARTES', 'MIERCOLES', 'JUEVES',
                                'VIERNES', 'SABADO', 'DOMINGO']
                        dia_semana = wmap[date_obj.weekday()]
                    except Exception:
                        dia_semana = "OTRO"

                    ruta  = ws.cell(row=r, column=6).value
                    lugar = ws.cell(row=r, column=10).value
                    monto = ws.cell(row=r, column=11).value

                    all_records.append({
                        "dia_semana": dia_semana,
                        "ruta":  str(ruta).strip()  if ruta  else "",
                        "lugar": str(lugar).strip() if lugar else "",
                        "monto": float(monto) if monto is not None else 0.0
                    })

            worker_wb.close()

            if not vendor_name:
                execution_logs.append(f"Archivo '{file.filename}' ignorado: sin nombre de trabajador.")
                continue
            if not banco_name or banco_name not in BANCOS_VALIDOS:
                execution_logs.append(
                    f"Archivo '{file.filename}' ignorado: banco '{banco_name}' no válido."
                )
                continue
            if not all_records:
                execution_logs.append(f"Archivo '{file.filename}' ignorado: sin registros de viaje.")
                continue

            # ── Seleccionar hoja del banco ──
            target_sheet_name = None
            for sname in master_wb.sheetnames:
                if sname.upper() == banco_name:
                    target_sheet_name = sname
                    break

            if not target_sheet_name:
                # Crear nueva hoja para el banco
                target_sheet_name = banco_name
                new_ws = master_wb.create_sheet(title=target_sheet_name)
                # Cabecera básica
                new_ws.merge_cells('B1:G1')
                new_ws['B1'] = f'TABLA DE PASAJES  - {banco_name}'
                for col, hdr in enumerate(['VENDEDOR', 'DIA ', 'RUTA', 'LUGAR', 'MONTO', 'TOTAL'], start=2):
                    new_ws.cell(row=2, column=col, value=hdr)
                # Fila TOTAL inicial
                new_ws.merge_cells(f'B3:F3')
                new_ws['B3'] = 'TOTAL'

            sheet = master_wb[target_sheet_name]

            # ── Buscar bloque existente del vendedor ──
            # Columna B contiene nombre del vendedor (merged)
            start_row = None
            end_row   = None

            for r in range(3, sheet.max_row + 1):
                val_b = sheet.cell(row=r, column=2).value
                if val_b and str(val_b).strip().upper() == vendor_name.upper():
                    start_row = r
                    # Buscar fin del merge en col B
                    for rng in sheet.merged_cells.ranges:
                        mc, mr, xc, xr = rng.bounds
                        if mc == 2 and mr == r:
                            end_row = xr
                            break
                    if end_row is None:
                        end_row = r
                    break

            # ── Eliminar bloque anterior si existe ──
            if start_row is not None:
                # Quitar merges del bloque
                to_remove = [
                    rng for rng in list(sheet.merged_cells.ranges)
                    if rng.bounds[1] >= start_row and rng.bounds[3] <= end_row
                ]
                for rng in to_remove:
                    sheet.merged_cells.remove(rng)
                rows_to_del = end_row - start_row + 1
                sheet.delete_rows(start_row, rows_to_del)
                insert_row = start_row
            else:
                # Insertar antes de la fila TOTAL
                total_row = None
                for r in range(3, sheet.max_row + 1):
                    v = sheet.cell(row=r, column=2).value
                    if v and str(v).strip().upper() == 'TOTAL':
                        total_row = r
                        break
                insert_row = total_row if total_row else sheet.max_row + 1

            # ── Insertar BLOCK_SIZE filas ──
            N = BLOCK_SIZE
            sheet.insert_rows(insert_row, N)

            # Estilos de referencia
            ref_font   = make_ref_font()
            ref_border = make_thin_border()

            # Escribir vendedor en col B (merge)
            c_vend = sheet.cell(row=insert_row, column=2, value=vendor_name)
            c_vend.font      = Font(name="Calibri", size=10, bold=True)
            c_vend.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            c_vend.border    = copy.copy(ref_border)
            sheet.merge_cells(
                start_row=insert_row, start_column=2,
                end_row=insert_row + N - 1, end_column=2
            )

            # Escribir total del vendedor en col G (merge)
            g_start = f"F{insert_row}"
            g_end   = f"F{insert_row + N - 1}"
            total_formula = f"=SUM(F{insert_row}:F{insert_row + N - 1})"
            c_total = sheet.cell(row=insert_row, column=7, value=total_formula)
            c_total.font         = copy.copy(ref_font)
            c_total.alignment    = Alignment(horizontal="right", vertical="center")
            c_total.number_format = '#,##0.00'
            c_total.border       = copy.copy(ref_border)
            sheet.merge_cells(
                start_row=insert_row, start_column=7,
                end_row=insert_row + N - 1, end_column=7
            )

            # Escribir registros (hasta BLOCK_SIZE filas)
            for idx in range(N):
                curr = insert_row + idx
                # Bordear col B y G
                for col in [2, 7]:
                    sheet.cell(row=curr, column=col).border = copy.copy(ref_border)

                if idx < len(all_records):
                    rec = all_records[idx]
                    c_dia   = sheet.cell(row=curr, column=3, value=rec['dia_semana'])
                    c_ruta  = sheet.cell(row=curr, column=4, value=rec['ruta'])
                    c_lugar = sheet.cell(row=curr, column=5, value=rec['lugar'])
                    c_monto = sheet.cell(row=curr, column=6, value=rec['monto'])

                    c_dia.alignment   = Alignment(horizontal="center", vertical="center")
                    c_ruta.alignment  = Alignment(horizontal="left",   vertical="center")
                    c_lugar.alignment = Alignment(horizontal="center", vertical="center")
                    c_monto.alignment = Alignment(horizontal="right",  vertical="center")
                    c_monto.number_format = '#,##0.00'

                    for col in [3, 4, 5, 6]:
                        sheet.cell(row=curr, column=col).font   = copy.copy(ref_font)
                        sheet.cell(row=curr, column=col).border = copy.copy(ref_border)
                else:
                    # Filas vacías con borde
                    for col in [3, 4, 5, 6]:
                        cell = sheet.cell(row=curr, column=col)
                        cell.value  = None
                        cell.border = copy.copy(ref_border)

            # ── Actualizar fórmula TOTAL global de la hoja ──
            grand_total_row = None
            for r in range(3, sheet.max_row + 1):
                v = sheet.cell(row=r, column=2).value
                if v and str(v).strip().upper() == 'TOTAL':
                    grand_total_row = r
                    break
            
            if not grand_total_row:
                # Crear fila de TOTAL al final si no existe
                grand_total_row = sheet.max_row + 1
                sheet.cell(row=grand_total_row, column=2, value="TOTAL")
                sheet.merge_cells(start_row=grand_total_row, start_column=2, end_row=grand_total_row, end_column=6)
                # Bordear la fila de TOTAL
                for col in range(2, 7):
                    sheet.cell(row=grand_total_row, column=col).border = copy.copy(ref_border)

            sheet.cell(
                row=grand_total_row, column=7,
                value=f"=SUM(G3:G{grand_total_row - 1})"
            )
            sheet.cell(row=grand_total_row, column=7).font = Font(name="Calibri", size=10, bold=True)
            sheet.cell(row=grand_total_row, column=7).border = copy.copy(ref_border)
            sheet.cell(row=grand_total_row, column=7).alignment = Alignment(horizontal="right", vertical="center")

            execution_logs.append(
                f"✓ {vendor_name} → hoja '{target_sheet_name}': "
                f"{min(len(all_records), N)} filas actualizadas."
            )
            consolidated_count += 1

        master_wb.save(MASTER_FILE)
        master_wb.close()

        return app.response_class(
            response=json.dumps({
                "status":             "success",
                "consolidated_count": consolidated_count,
                "logs":               [log.encode('ascii', 'replace').decode('ascii') for log in execution_logs]
            }, ensure_ascii=True),
            mimetype='application/json'
        )
    except Exception as e:
        print("Error en consolidación:", e)
        return app.response_class(
            response=json.dumps({"status": "error", "message": str(e)}, ensure_ascii=True),
            mimetype='application/json'
        ), 500

if __name__ == '__main__':
    # Abrir navegador automáticamente solo si está compilado o si es el subproceso principal de Werkzeug en desarrollo
    is_frozen = getattr(sys, 'frozen', False)
    debug_mode = not is_frozen

    if not debug_mode or os.environ.get('WERKZEUG_RUN_MAIN') == 'true':
        import webbrowser
        from threading import Timer
        def open_browser():
            webbrowser.open_new("http://127.0.0.1:5000")
        Timer(1.2, open_browser).start()

    app.run(host='0.0.0.0', port=5000, debug=debug_mode)

