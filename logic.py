import openpyxl
from openpyxl.styles import Font, Border, Alignment, PatternFill, Side
import os
import json
import io
import copy
from datetime import datetime, date
import sys

def resource_path(relative_path):
    if getattr(sys, 'frozen', False):
        return os.path.join(sys._MEIPASS, relative_path)
    return os.path.join(os.path.abspath("."), relative_path)

base_dir = os.path.dirname(sys.executable) if getattr(sys, 'frozen', False) else os.path.dirname(os.path.abspath(__file__))
DNI_FILE = os.path.join(base_dir, 'vendors_dni.json')
MASTER_FILE = os.path.join(base_dir, '2. PASAJES FFVV JULIO 2026 MODELO.xlsx')
TEMPLATE_FILE = os.path.join(base_dir, '1. Formato Planilla de movilidad - por trabajador.xlsx')

BANCOS_VALIDOS = ['BBVA', 'BCP', 'SCOTIABANK', 'FINANCIERO']

def load_dnis():
    if os.path.exists(DNI_FILE):
        try:
            with open(DNI_FILE, 'r', encoding='utf-8') as f:
                return json.load(f)
        except Exception:
            pass
    return {}

def save_vendor_info(vendor_name, dni, banco):
    dnis = load_dnis()
    dnis[vendor_name.strip()] = {
        'dni': dni.strip(),
        'banco': banco.strip().upper()
    }
    try:
        with open(DNI_FILE, 'w', encoding='utf-8') as f:
            json.dump(dnis, f, ensure_ascii=False, indent=4)
    except Exception as e:
        print("Error saving vendor JSON:", e)

def extract_vendors_and_defaults():
    vendors = {}
    if not os.path.exists(MASTER_FILE):
        return vendors
    try:
        wb = openpyxl.load_workbook(MASTER_FILE, data_only=True)
        for sheet_name in wb.sheetnames:
            if sheet_name.upper() not in [b.upper() for b in BANCOS_VALIDOS]:
                continue
            sheet = wb[sheet_name]
            banco_nombre = sheet_name.upper()
            current_vendor = None
            for r in range(3, sheet.max_row + 1):
                val_b = sheet.cell(row=r, column=2).value
                val_c = sheet.cell(row=r, column=3).value
                val_f = sheet.cell(row=r, column=6).value
                if val_b and str(val_b).strip().upper() == 'TOTAL':
                    break
                if val_b is not None and str(val_b).strip():
                    current_vendor = str(val_b).strip()
                    if current_vendor not in vendors:
                        vendors[current_vendor] = {"banco": banco_nombre, "defaults": {}}
                if current_vendor:
                    dia = str(val_c).strip().upper() if val_c else ""
                    ruta = sheet.cell(row=r, column=4).value
                    lugar = sheet.cell(row=r, column=5).value
                    monto = sheet.cell(row=r, column=6).value
                    if dia in ['LUNES', 'MARTES', 'MIERCOLES', 'JUEVES', 'VIERNES', 'SABADO']:
                        if dia not in vendors[current_vendor]["defaults"]:
                            vendors[current_vendor]["defaults"][dia] = {
                                "ruta": str(ruta).strip() if ruta else "",
                                "lugar": str(lugar).strip() if lugar else "",
                                "monto": float(monto) if monto is not None else 0.0
                            }
        wb.close()
    except Exception as e:
        print("Error extracting vendor defaults:", e)
    return vendors

def generate_planilla_excel(vendor, dni, banco, periodo_str, fecha_emision_s, planilla_nro, rows):
    if not os.path.exists(TEMPLATE_FILE):
        raise FileNotFoundError("Archivo plantilla no encontrado")
    
    save_vendor_info(vendor, dni, banco)
    wb = openpyxl.load_workbook(TEMPLATE_FILE)
    template_sheet = wb['Planilla Movilidad']

    p_yr, p_mo = map(int, periodo_str.split('-'))
    periodo_date = date(p_yr, p_mo, 1)
    fecha_emision = datetime.strptime(fecha_emision_s, "%Y-%m-%d").date()

    chunk_size = 14
    row_chunks = [rows[i:i + chunk_size] for i in range(0, len(rows), chunk_size)]

    for p_idx, chunk in enumerate(row_chunks):
        if p_idx == 0:
            sheet = template_sheet
            sheet.title = f"Planilla Pg {p_idx + 1}"
        else:
            sheet = wb.copy_worksheet(template_sheet)
            sheet.title = f"Planilla Pg {p_idx + 1}"

        sheet['F13'] = vendor
        sheet['F14'] = dni
        sheet['K11'] = banco
        sheet['D8']  = periodo_date
        sheet['D8'].number_format = 'mmm-yy'
        sheet['K5']  = fecha_emision
        sheet['K5'].number_format = 'dd/mm/yyyy'
        sheet['K1']  = f"{planilla_nro + p_idx:02d}"

        for r_idx in range(chunk_size):
            excel_row = 18 + r_idx
            if r_idx < len(chunk):
                rec = chunk[r_idx]
                sheet.cell(row=excel_row, column=1,  value=int(rec['dia']))
                sheet.cell(row=excel_row, column=2,  value=int(rec['mes']))
                sheet.cell(row=excel_row, column=3,  value=int(rec['ano']))
                sheet.cell(row=excel_row, column=4,  value=rec['motivo'])
                sheet.cell(row=excel_row, column=6,  value=rec['ruta'])
                sheet.cell(row=excel_row, column=10, value=rec['lugar'])
                sheet.cell(row=excel_row, column=11, value=float(rec['monto']))
            else:
                for col in [1, 2, 3, 4, 6, 10, 11]:
                    sheet.cell(row=excel_row, column=col, value=None)
        sheet['K32'] = "=SUM(K18:K31)"

    filename = f"Planilla_Movilidad_{vendor.replace(' ', '_')}_{banco}_{periodo_str.replace('-', '_')}.xlsx"
    downloads_path = os.path.join(os.path.expanduser('~'), 'Downloads')
    
    # Check if Android environment (where HOME is not well defined)
    if 'ANDROID_DATA' in os.environ:
        filepath = os.path.join('/storage/emulated/0/Download', filename)
    else:
        filepath = os.path.join(downloads_path, filename)
        
    wb.save(filepath)
    wb.close()
    return filepath


def consolidate_planillas_excel(file_paths: list):
    """
    Consolida una lista de archivos .xlsx de planillas individuales
    en el archivo maestro. Guarda el resultado en Descargas.
    Retorna (filepath, logs).
    """
    if not os.path.exists(MASTER_FILE):
        raise FileNotFoundError(f"Archivo maestro no encontrado: {MASTER_FILE}")

    master_wb = openpyxl.load_workbook(MASTER_FILE)

    # Limpiar datos previos de vendedores en hojas de banco
    for sheet_name in master_wb.sheetnames:
        if sheet_name.upper() not in [b.upper() for b in BANCOS_VALIDOS]:
            continue
        sheet = master_wb[sheet_name]
        total_row = None
        for r in range(3, sheet.max_row + 1):
            val = sheet.cell(row=r, column=2).value
            if val and str(val).strip().upper() == 'TOTAL':
                total_row = r
                break
        if total_row and total_row > 3:
            to_remove = [rng for rng in list(sheet.merged_cells.ranges)
                         if rng.bounds[1] >= 3 and rng.bounds[3] < total_row]
            for rng in to_remove:
                sheet.merged_cells.remove(rng)
            sheet.delete_rows(3, total_row - 3)

    consolidated_count = 0
    execution_logs = []
    BLOCK_SIZE = 26

    for path in file_paths:
        worker_wb = openpyxl.load_workbook(path, data_only=True)
        vendor_name = None
        banco_name  = None
        all_records = []

        for ws in worker_wb.worksheets:
            v_name  = ws['F13'].value
            v_banco = ws['K11'].value
            if v_name:
                vendor_name = str(v_name).strip()
            if v_banco:
                banco_name = str(v_banco).strip().upper()

            for r in range(18, 32):
                dia_val = ws.cell(row=r, column=1).value
                if dia_val is None:
                    continue
                try:
                    d  = int(dia_val)
                    mo = int(ws.cell(row=r, column=2).value or 1)
                    yr = int(ws.cell(row=r, column=3).value or 2026)
                    date_obj = date(yr, mo, d)
                    wmap = ['LUNES', 'MARTES', 'MIERCOLES', 'JUEVES', 'VIERNES', 'SABADO', 'DOMINGO']
                    dia_semana = wmap[date_obj.weekday()]
                except Exception:
                    dia_semana = "OTRO"
                ruta  = ws.cell(r, 6).value
                lugar = ws.cell(r, 10).value
                monto = ws.cell(r, 11).value
                all_records.append({
                    "dia_semana": dia_semana,
                    "ruta":  str(ruta).strip()  if ruta  else "",
                    "lugar": str(lugar).strip() if lugar else "",
                    "monto": float(monto) if monto is not None else 0.0
                })
        worker_wb.close()

        if not vendor_name:
            execution_logs.append(f"'{os.path.basename(path)}' ignorado: sin nombre de trabajador.")
            continue
        if not banco_name or banco_name not in BANCOS_VALIDOS:
            execution_logs.append(f"'{os.path.basename(path)}' ignorado: banco '{banco_name}' no válido.")
            continue
        if not all_records:
            execution_logs.append(f"'{os.path.basename(path)}' ignorado: sin registros.")
            continue

        target_sheet_name = next((s for s in master_wb.sheetnames if s.upper() == banco_name), None)
        if not target_sheet_name:
            target_sheet_name = banco_name
            new_ws = master_wb.create_sheet(title=target_sheet_name)
            new_ws.merge_cells('B1:G1')
            new_ws['B1'] = f'TABLA DE PASAJES - {banco_name}'
            for col, hdr in enumerate(['VENDEDOR', 'DIA', 'RUTA', 'LUGAR', 'MONTO', 'TOTAL'], start=2):
                new_ws.cell(row=2, column=col, value=hdr)
            new_ws.merge_cells('B3:F3')
            new_ws['B3'] = 'TOTAL'

        sheet = master_wb[target_sheet_name]

        total_row = next((r for r in range(3, sheet.max_row + 1)
                          if str(sheet.cell(r, 2).value or '').strip().upper() == 'TOTAL'), None)
        insert_row = total_row if total_row else sheet.max_row + 1

        N = BLOCK_SIZE
        sheet.insert_rows(insert_row, N)

        ref_font   = Font(name="Calibri", size=10)
        ref_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                             top=Side(style='thin'), bottom=Side(style='thin'))

        c_vend = sheet.cell(insert_row, 2, value=vendor_name)
        c_vend.font      = Font(name="Calibri", size=10, bold=True)
        c_vend.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c_vend.border    = copy.copy(ref_border)
        sheet.merge_cells(start_row=insert_row, start_column=2,
                          end_row=insert_row + N - 1, end_column=2)

        c_total = sheet.cell(insert_row, 7, value=f"=SUM(F{insert_row}:F{insert_row + N - 1})")
        c_total.font          = copy.copy(ref_font)
        c_total.alignment     = Alignment(horizontal="right", vertical="center")
        c_total.number_format = '#,##0.00'
        c_total.border        = copy.copy(ref_border)
        sheet.merge_cells(start_row=insert_row, start_column=7,
                          end_row=insert_row + N - 1, end_column=7)

        for idx in range(N):
            curr = insert_row + idx
            for col in [2, 7]:
                sheet.cell(curr, col).border = copy.copy(ref_border)
            if idx < len(all_records):
                rec = all_records[idx]
                c_d = sheet.cell(curr, 3, value=rec['dia_semana'])
                c_r = sheet.cell(curr, 4, value=rec['ruta'])
                c_l = sheet.cell(curr, 5, value=rec['lugar'])
                c_m = sheet.cell(curr, 6, value=rec['monto'])
                c_d.alignment = Alignment(horizontal="center", vertical="center")
                c_r.alignment = Alignment(horizontal="left",   vertical="center")
                c_l.alignment = Alignment(horizontal="center", vertical="center")
                c_m.alignment = Alignment(horizontal="right",  vertical="center")
                c_m.number_format = '#,##0.00'
                for col in [3, 4, 5, 6]:
                    sheet.cell(curr, col).font   = copy.copy(ref_font)
                    sheet.cell(curr, col).border = copy.copy(ref_border)
            else:
                for col in [3, 4, 5, 6]:
                    cell = sheet.cell(curr, col)
                    cell.value  = None
                    cell.border = copy.copy(ref_border)

        grand_total_row = next(
            (r for r in range(3, sheet.max_row + 1)
             if str(sheet.cell(r, 2).value or '').strip().upper() == 'TOTAL'), None)
        if not grand_total_row:
            grand_total_row = sheet.max_row + 1
            sheet.cell(grand_total_row, 2, value="TOTAL")
            sheet.merge_cells(start_row=grand_total_row, start_column=2,
                              end_row=grand_total_row, end_column=6)
        c_gt = sheet.cell(grand_total_row, 7, value=f"=SUM(G3:G{grand_total_row - 1})")
        c_gt.font      = Font(name="Calibri", size=10, bold=True)
        c_gt.border    = copy.copy(ref_border)
        c_gt.alignment = Alignment(horizontal="right", vertical="center")

        execution_logs.append(f"✔ {vendor_name} → hoja '{target_sheet_name}': {min(len(all_records), N)} filas.")
        consolidated_count += 1

    if consolidated_count == 0:
        master_wb.close()
        raise ValueError("No se pudo consolidar ningún archivo. Revisa los detalles.")

    filename = f"Consolidado_Pasajes_FFVV_{datetime.now().strftime('%Y_%m')}.xlsx"
    downloads_path = os.path.join(os.path.expanduser('~'), 'Downloads')
    if 'ANDROID_DATA' in os.environ:
        filepath = os.path.join('/storage/emulated/0/Download', filename)
    else:
        filepath = os.path.join(downloads_path, filename)

    master_wb.save(filepath)
    master_wb.close()
    return filepath, execution_logs
